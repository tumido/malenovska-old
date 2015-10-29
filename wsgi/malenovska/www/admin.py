from django.contrib import admin
from django.utils import timezone
from django.http import HttpResponse
from .models import Race, Player, Legend, News, AboutWidget, DateOptions, TextOptions, MapPoints, Harmonogram, ExtraFiles
from django.contrib.auth.admin import UserAdmin

def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registrace.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Registrovaní hráči"

    row_num = 0

    columns = [
        (u"Rasa", 20),
        (u"Přezdívka", 20),
        (u"Jméno", 20),
        (u"Příjmení", 20),
        (u"Věk", 20)
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset.order_by('race', 'nick', 'surname'):
        row_num += 1
        row = [obj.race.name,
               obj.nick,
               obj.name,
               obj.surname,
               obj.age]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Exportovat jako XLSX (Excel)"

@admin.register(Player)
class PlayerAdmin(admin.ModelAdmin):
    actions = [export_xlsx]
    fieldsets = [
        (None,               {'fields': [('nick', 'name', 'surname'), ('race', 'age')]}),
        ('Detaily', {'fields': ['email', 'date', 'ip'], 'classes': ['collapse']}),
    ]
    list_display = ['race', 'nick', 'name', 'surname', 'age']
    ordering = ('race', 'nick', 'surname')

    def get_form(self, request, obj=None, **kwargs):
        form = super(PlayerAdmin, self).get_form(request, obj, **kwargs)
        form.base_fields['date'].initial = timezone.now()
        form.base_fields['ip'].initial = '0.0.0.0'
        form.base_fields['email'].initial = 'noexist@malenovska.cz'
        return form


@admin.register(Race)
class RaceAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('name', 'active', 'limit')]}),
        ('Popis', {'fields': ['icon', 'fraction', 'description'], 'classes': ['collapse']}),
    ]
    list_display = ['name', 'active', 'limit']


@admin.register(Legend)
class LegendAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('name', 'race'), 'text']}),
    ]

@admin.register(ExtraFiles)
class FilesAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('name', 'tooltip', 'filefield')]}),
    ]

@admin.register(DateOptions)
class DateOptionsAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('name', 'date')]}),
        ('Skryto', {'fields': ['identifier'], 'classes': ['collapse']}),
    ]
    list_display = ['name', 'date']


@admin.register(TextOptions)
class TextOptionsAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('name', 'text')]}),
        ('Skryto', {'fields': ['identifier'], 'classes': ['collapse']}),
    ]
    list_display = ['name', 'text']


@admin.register(MapPoints)
class MapAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('title', 'lat', 'long')]}),
    ]


@admin.register(AboutWidget)
class AboutWidgetAdmin(admin.ModelAdmin):
    list_display = ['name', 'identifier']
    fieldsets = [
        (None,               {'fields': ['name', 'text']}),
        ('Skryto', {'fields': ['identifier'], 'classes': ['collapse']}),
    ]


@admin.register(Harmonogram)
class HarmonogramAdmin(admin.ModelAdmin):
    fieldsets = [
        (None,               {'fields': [('program', 'time_start', 'time_end')]}),
    ]
    list_display = ['time_start', 'time_end', 'program']

admin.site.register(News)

def roles(self):
    #short_name = unicode # function to get group name
    short_name = lambda x: x.name # first letter of a group
    p = sorted([u"<a title='%s'>%s</a>" % (x, short_name(x)) for x in self.groups.all()])
    if self.user_permissions.count(): p += ['+']
    value = ', '.join(p)
    return value
roles.allow_tags = True
roles.short_description = u'Groups'

UserAdmin.list_display += (roles,)
